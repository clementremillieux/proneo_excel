"""_summary_"""
import os

from datetime import datetime


import time
from typing import List, Optional

import openpyxl

from openpyxl.worksheet.worksheet import Worksheet

import xlwings as xw

from config.logger_config import logger

from modules.performances.time_counter import time_execution


def add_xlwings_conf_sheet(file_path: str):
    """
    Open an Excel file using openpyxl, create a hidden sheet named 'xlwings.conf',
    and add the configuration settings.

    Args:
        file_path (str): Path to the Excel file where the hidden sheet should be added.
    """

    try:
        if "onedrive" not in file_path.lower():
            return
        
        workbook = openpyxl.load_workbook(file_path, keep_vba=True)

        one_drive_root = os.path.dirname(file_path)

        while True:
            logger.warning(one_drive_root.rsplit("/",1)[-1])
            if "onedrive" not in one_drive_root.rsplit("/",1)[-1].lower():
                one_drive_root = one_drive_root.rsplit("/",1)[0]
            else:
                break

            if len(one_drive_root.rsplit("/",1)) < 2:
                break

        xlwings_conf_sheet: Worksheet = workbook.create_sheet(
            title='xlwings.conf')

        xlwings_conf_sheet.sheet_state = 'hidden'

        logger.warning(one_drive_root)

        config_data = [["ONEDRIVE_CONSUMER_MAC", one_drive_root],
                       ["ONEDRIVE_COMMERCIAL_MAC", one_drive_root],
                       ["SHAREPOINT_MAC", one_drive_root],
                       ["ONEDRIVE_CONSUMER_WIN", one_drive_root],
                       ["ONEDRIVE_COMMERCIAL_WIN", one_drive_root],
                       ["SHAREPOINT_WIN", one_drive_root]]

        for row_index, (key, value) in enumerate(config_data, start=1):
            xlwings_conf_sheet.cell(row=row_index, column=1, value=key)

            xlwings_conf_sheet.cell(row=row_index, column=2, value=value)

        workbook.save(file_path)

        logger.info(
            "xlwings.conf [%s] sheet has been added and configured successfully in '%s'.",
           one_drive_root, file_path)

        workbook.close()

        time.sleep(10)

    except Exception as e:
        logger.info("An error occurred while modifying the Excel file: %s", e)


class ExcelHandler:
    """Handles interaction with Excel files using xlwings."""

    def __init__(self) -> None:
        self.excel_abs_path: Optional[str] = None

        self.app: Optional[xw.App] = None

        self.wb: Optional[xw.Book] = None

    @time_execution
    def load_excel(self, excel_abs_path: str) -> None:
        """Loads an Excel file with xlwings."""
        try:
            self.excel_abs_path = excel_abs_path

            self.app = xw.apps.active

            if self.app is None:
                self.app = xw.App(visible=True)

            for book in self.app.books:

                if book.fullname == excel_abs_path:
                    self.wb = book

                    break
            else:

                self.wb = self.app.books.open(excel_abs_path)

        except Exception as e:
            logger.error("Error opening Excel file with xlwings: %s",
                         e,
                         exc_info=True)

    def read_cell_value(self, sheet_name: str, cell_address: str) -> str:
        """Reads the value from a specific cell."""
        try:
            sheet = self.wb.sheets[sheet_name]

            cell_value = sheet.range(cell_address).value

            return cell_value

        except Exception as e:
            logger.error("Error reading cell: %s", e)

            return ""

    def read_cell_date_value(self, sheet_name: str,
                             cell_address: str) -> datetime:
        """Reads the value of a specific cell and returns it as a datetime."""
        try:
            sheet = self.wb.sheets[sheet_name]

            cell_value: datetime = sheet.range(cell_address).value

            return cell_value

        except Exception as e:
            logger.error("Error reading cell date: %s", e)

            return datetime.min

    def get_checkbox_state(self, sheet_name: str, cell_address: str) -> bool:
        """Checks if a checkbox is checked."""
        try:
            sheet = self.wb.sheets[sheet_name]

            checkbox_value: bool = sheet.range(cell_address).value

            return checkbox_value

        except Exception as e:
            logger.error("Error reading checkbox state: %s", e)
            return False

    def get_all_sheets(self) -> List[str]:
        """Gets all sheet names in the workbook."""
        try:
            sheet_names = [sheet.name for sheet in self.wb.sheets]

            return sheet_names

        except Exception as e:
            logger.error("Error getting sheet names: %s", e)
            return []

    def go_to_sheet_and_cell(self, sheet_name: str, cell_address: str) -> None:
        """Activates a sheet and selects a specific cell."""

        try:
            self.load_excel(excel_abs_path=self.excel_abs_path)

            sheet = self.wb.sheets[sheet_name]

            sheet.activate()

            sheet.range(cell_address).select()

        except Exception as e:
            logger.error("Error navigating to %s in sheet %s: %s",
                         cell_address, sheet_name, e)

    def is_merged(self, sheet_name: str, cell_address: str) -> bool:
        """Checks if a cell is part of a merged range."""
        try:
            sheet = self.wb.sheets[sheet_name]

            cell = sheet.range(cell_address)

            if cell.merge_area.count > 1:
                return cell.address != cell.merge_area[0].address

            return False

        except Exception as e:
            logger.error(
                f"Error checking if cell {cell_address} is merged: {e}")
            return False

    def is_column_hidden(self, sheet_name: str, cell_address: str) -> bool:
        """Checks if the column is hidden cross-platform (Windows and Mac)."""
        try:
            sheet = self.wb.sheets[sheet_name]

            column = cell_address[0]

            column_width = sheet.range(f'{column}:{column}').column_width

            return column_width == 0

        except Exception as e:
            logger.error(f"Error reading column hidden state: {e}")
            return False

    def is_row_hidden(self, sheet_name: str, cell_address: str) -> bool:
        """Checks if the row is hidden cross-platform (Windows and Mac)."""
        try:
            sheet = self.wb.sheets[sheet_name]

            row = int(cell_address[1:])

            row_height = sheet.range(f'{row}:{row}').row_height

            return row_height == 0

        except Exception as e:
            logger.error(f"Error reading row hidden state: {e}")
            return False
