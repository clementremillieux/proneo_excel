from openpyxl import load_workbook
import inspect


def explore_worksheet(ws):
    print(f"\nExploring worksheet: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")
    print(f"Max Row: {ws.max_row}")
    print(f"Max Column: {ws.max_column}")

    # Explore worksheet attributes
    print("\nWorksheet attributes:")
    for attr in dir(ws):
        if not attr.startswith("_"):  # Exclude private attributes
            try:
                value = getattr(ws, attr)
                if not callable(value):
                    print(f"{attr}: {value}")
            except Exception as e:
                print(f"{attr}: Unable to retrieve (Error: {str(e)})")

    # Explore cells
    print("\nExploring cells for potential checkbox indicators:")
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
        for cell in row:
            if cell.value is not None:
                print(f"Cell {cell.coordinate}:")
                print(f"  Value: {cell.value}")
                print(f"  Data type: {cell.data_type}")

                # Check for hyperlinks
                if cell.hyperlink:
                    print(f"  Hyperlink: {cell.hyperlink}")

                # Check for comments
                if cell.comment:
                    print(f"  Comment: {cell.comment.text}")

    # Check for named ranges
    # print("\nNamed ranges:")
    # for named_range in ws.defined_names.definedName:
    #     print(f"Name: {named_range.name}, Reference: {named_range.attr_text}")

    # # Check for conditional formatting
    # print("\nConditional Formatting:")
    # if ws.conditional_formatting:
    #     for cf in ws.conditional_formatting:
    #         print(f"Range: {cf.cells.ranges}, Rules: {cf.rules}")


def explore_workbook(workbook_path):
    wb = load_workbook(workbook_path, keep_vba=True, data_only=True)

    print("Workbook Properties:")
    for attr in dir(wb.properties):
        if not attr.startswith("_"):
            try:
                value = getattr(wb.properties, attr)
                if not callable(value):
                    print(f"{attr}: {value}")
            except Exception as e:
                print(f"{attr}: Unable to retrieve (Error: {str(e)})")

    print("\nExploring VBA Project:")
    if wb.vba_archive:
        for item in wb.vba_archive.namelist():
            print(item)
    else:
        print("No VBA archive found")

    for ws in wb.worksheets[:2]:
        explore_worksheet(ws)


# Usage
workbook_path = "/Users/remillieux/Documents/Proneo/logiciel/data/Plan et Rapport d'audit certification V32.xlsm"
explore_workbook(workbook_path)
